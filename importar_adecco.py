"""
importar_adecco.py
==================
Procesador de importación para archivos FORMATO_NOVEDADES de ADECCO / Collective Mining.
Integrado con la base de datos MineDax (SQL Server).

Dos funcionalidades:
  1. Importar empleados (Maestro Original → GN_TERCE + GN_FUNCI)
  2. Importar novedades (Ocasionales / Fijas / Ausentismos / Cambios e Ingresos → NO_NOVED + subtablas)

Archivos protegidos con contraseña:
  La contraseña se lee desde el archivo .env (variable NOMINA_XLSX_PWD).
  El archivo en disco NUNCA se modifica ni desprotege — se desencripta en memoria.

Uso:
    python importar_adecco.py <ruta_excel>

Requiere (instalar una vez):
    pip install openpyxl pyodbc msoffcrypto-tool
"""

import sys
import os
import io
import re
import unicodedata
import pyodbc
import openpyxl
import msoffcrypto
from datetime import datetime, date
from typing import Optional
from pathlib import Path

# ─────────────────────────────────────────────────────────────
# CARGAR .env (busca en la carpeta del script y sus padres)
# ─────────────────────────────────────────────────────────────
def _load_dotenv():
    """Carga variables del .env más cercano sin depender de python-dotenv."""
    ruta = Path(__file__).resolve()
    for parent in [ruta.parent] + list(ruta.parents):
        env_file = parent / ".env"
        if env_file.exists():
            with open(env_file, encoding="utf-8") as f:
                for linea in f:
                    linea = linea.strip()
                    if not linea or linea.startswith("#") or "=" not in linea:
                        continue
                    clave, _, valor = linea.partition("=")
                    clave = clave.strip()
                    valor = valor.split("#")[0].strip()  # ignorar comentarios inline
                    if clave and clave not in os.environ:
                        os.environ[clave] = valor
            break

_load_dotenv()

# ─────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE CONEXIÓN
# ─────────────────────────────────────────────────────────────
DB_SERVER   = os.environ.get("SERVER", r"CM-ITD-P-05\SQLEXPRESS")
DB_NAME     = os.environ.get("DATABASE", "MineDax")
DB_UID      = os.environ.get("UID", "")
DB_PWD      = os.environ.get("PWD", "")
COD_EMPR    = 1             # Empresa activa

# Contraseña para abrir archivos Excel de nómina protegidos
NOMINA_XLSX_PWD = os.environ.get("NOMINA_XLSX_PWD", "")

def get_conn():
    if DB_UID and DB_PWD:
        conn_str = (
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={DB_SERVER};DATABASE={DB_NAME};"
            f"UID={DB_UID};PWD={DB_PWD};"
        )
    else:
        conn_str = (
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={DB_SERVER};DATABASE={DB_NAME};"
            f"Trusted_Connection=yes;"
        )
    return pyodbc.connect(conn_str, autocommit=False)


# ─────────────────────────────────────────────────────────────
# ABRIR EXCEL (con o sin contraseña, siempre en memoria)
# ─────────────────────────────────────────────────────────────
def abrir_excel(ruta: str) -> openpyxl.Workbook:
    """
    Abre un archivo Excel, desencriptándolo en memoria si está protegido.
    El archivo original en disco no se modifica en ningún caso.
    Si la contraseña falla, lanza un error descriptivo.
    """
    with open(ruta, "rb") as f:
        raw = f.read()

    # Detectar si está cifrado
    buf_in = io.BytesIO(raw)
    try:
        office = msoffcrypto.OfficeFile(buf_in)
        is_enc = office.is_encrypted()
    except Exception:
        is_enc = False

    if not is_enc:
        # Archivo sin contraseña — cargar directo
        return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    # Archivo cifrado — desencriptar en memoria
    if not NOMINA_XLSX_PWD:
        raise RuntimeError(
            "El archivo está protegido con contraseña pero NOMINA_XLSX_PWD "
            "no está definida en el archivo .env.\n"
            "Agrega la línea:  NOMINA_XLSX_PWD=<contraseña>  al archivo .env"
        )

    buf_in = io.BytesIO(raw)
    office = msoffcrypto.OfficeFile(buf_in)
    office.load_key(password=NOMINA_XLSX_PWD)

    buf_out = io.BytesIO()
    try:
        office.decrypt(buf_out)
    except Exception as e:
        raise RuntimeError(
            f"No se pudo desencriptar el archivo. "
            f"Verifica que NOMINA_XLSX_PWD sea correcta.\nDetalle: {e}"
        )

    buf_out.seek(0)
    return openpyxl.load_workbook(buf_out, data_only=True)


# ─────────────────────────────────────────────────────────────
# UTILIDADES GENERALES
# ─────────────────────────────────────────────────────────────

def normalize(s: str) -> str:
    """Elimina tildes y pone en mayúsculas para comparaciones fuzzy."""
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(s))
    ascii_ = nfkd.encode("ascii", "ignore").decode("ascii")
    return ascii_.upper().strip()


def clean(val, max_len: int = None) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "None", "#N/A", "N/A", "0000", "NA"):
        return None
    if max_len:
        s = s[:max_len]
    return s


def parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def split_nombre(nombre_completo: str):
    """
    Divide el nombre completo en formato ADECCO: APELLIDO1 APELLIDO2 NOMBRE1 NOMBRE2.
    Retorna: (APE_TERC, SEG_APEL, NOM_TERC, SEG_NOMB)
    """
    partes = nombre_completo.strip().upper().split()
    if len(partes) == 0:
        return "", "", "", ""
    if len(partes) == 1:
        return partes[0], "", "", ""
    if len(partes) == 2:
        # APE1 NOM1
        return partes[0], "", partes[1], ""
    if len(partes) == 3:
        # APE1 APE2 NOM1
        return partes[0], partes[1], partes[2], ""
    # 4+: APE1 APE2 NOM1 NOM2(+)
    return partes[0], partes[1], partes[2], " ".join(partes[3:])


def extraer_periodo_nombre(filename: str):
    """
    Extrae año y quincena del nombre del archivo.
    Ejemplos compatibles:
      '00-_FORMATO_NOVEDADES_CM_2Q_FEB_2026'  → (2026, 2, 2)
      '00- FORMATO NOVEDADES CM 2Q ABRIL 2026' → (2026, 4, 2)
    Toma los primeros 3 caracteres del mes para el lookup (FEB, ABR, ABRIL→ABR…)
    """
    meses = {
        "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
        "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12
    }
    base = os.path.basename(filename).upper()
    m = re.search(r'(\d)Q[\s_](\w+)[\s_](\d{4})', base)
    if m:
        qna  = int(m.group(1))
        mes  = meses.get(m.group(2)[:3], 0)
        anio = int(m.group(3))
        return anio, mes, qna
    return None, None, None


# ─────────────────────────────────────────────────────────────
# CARGAR TABLAS MAESTRAS EN MEMORIA
# ─────────────────────────────────────────────────────────────

def cargar_maestras(cur):
    """Retorna dict con todas las tablas maestras normalizadas para lookup."""
    m = {}

    # MAE_TPDOC
    cur.execute("SELECT COD_TPDOC, NOM_TPDOC, COD_ABREV FROM MAE_TPDOC WHERE ACT_ESTA='A'")
    m["tpdoc_por_abrev"] = {normalize(r[2]): r[0] for r in cur.fetchall()}
    m["tpdoc_por_nombre"] = {normalize(r[1]): r[0] for r in []}  # rebuilt below
    cur.execute("SELECT COD_TPDOC, NOM_TPDOC, COD_ABREV FROM MAE_TPDOC WHERE ACT_ESTA='A'")
    rows = cur.fetchall()
    m["tpdoc_por_abrev"]  = {normalize(r[2]): r[0] for r in rows}
    m["tpdoc_por_nombre"] = {normalize(r[1]): r[0] for r in rows}

    # MAE_MUNI
    cur.execute("SELECT COD_MUNI, NOM_MUNI FROM MAE_MUNI WHERE ACT_ESTA='A'")
    rows = cur.fetchall()
    m["muni_por_nombre"] = {normalize(r[1]): r[0] for r in rows}

    # MAE_GRSAN
    cur.execute("SELECT COD_GRSAN, COD_LETRA, COD_FCRH FROM MAE_GRSAN WHERE ACT_ESTA='A'")
    rows = cur.fetchall()
    m["grsan_por_clave"] = {
        (normalize(r[1].strip()), normalize(r[2].strip())): r[0]
        for r in rows
    }

    # MAE_ESTCIV
    cur.execute("SELECT COD_ESTCIV, NOM_ESTCIV FROM MAE_ESTCIV WHERE ACT_ESTA='A'")
    rows = cur.fetchall()
    m["estciv_por_nombre"] = {normalize(r[1]): r[0] for r in rows}

    # MAE_CARGO
    cur.execute("SELECT COD_CARGO, NOM_CARGO FROM MAE_CARGO WHERE ACT_ESTA='A'")
    rows = cur.fetchall()
    m["cargo_por_nombre"] = {normalize(r[1]): r[0] for r in rows}

    # MAE_CCOST (por nombre y por abrev)
    cur.execute("SELECT COD_CCOST, NOM_CCOST, COD_ABREV FROM MAE_CCOST WHERE ACT_ESTA='A'")
    rows = cur.fetchall()
    m["ccost_por_nombre"] = {normalize(r[1]): r[0] for r in rows}
    m["ccost_por_abrev"]  = {normalize(r[2]): r[0] for r in rows}

    # MAE_BANCO
    cur.execute("SELECT COD_BANCO, NOM_BANCO FROM MAE_BANCO WHERE ACT_ESTA='A'")
    rows = cur.fetchall()
    m["banco_por_nombre"] = {normalize(r[1]): r[0] for r in rows}

    # MAE_TPCTA
    cur.execute("SELECT COD_TPCTA, NOM_TPCTA, COD_ABREV FROM MAE_TPCTA WHERE ACT_ESTA='A'")
    rows = cur.fetchall()
    m["tpcta_por_nombre"] = {normalize(r[1]): r[0] for r in rows}
    m["tpcta_por_abrev"]  = {normalize(r[2]): r[0] for r in rows}

    # MAE_EPS (via GN_TERCE)
    cur.execute("""
        SELECT e.COD_EPS, t.NOM_COMP
        FROM MAE_EPS e LEFT JOIN GN_TERCE t ON e.COD_TERC = t.COD_TERC
        WHERE e.ACT_ESTA='A'
    """)
    rows = cur.fetchall()
    m["eps_por_nombre"] = {normalize(r[1] or ""): r[0] for r in rows if r[1]}

    # MAE_AFP (via GN_TERCE)
    cur.execute("""
        SELECT a.COD_AFP, t.NOM_COMP
        FROM MAE_AFP a LEFT JOIN GN_TERCE t ON a.COD_TERC = t.COD_TERC
        WHERE a.ACT_ESTA='A'
    """)
    rows = cur.fetchall()
    m["afp_por_nombre"] = {normalize(r[1] or ""): r[0] for r in rows if r[1]}

    # MAE_CCF (via GN_TERCE)
    cur.execute("""
        SELECT c.COD_CCF, t.NOM_COMP
        FROM MAE_CCF c LEFT JOIN GN_TERCE t ON c.COD_TERC = t.COD_TERC
        WHERE c.ACT_ESTA='A'
    """)
    rows = cur.fetchall()
    m["ccf_por_nombre"] = {normalize(r[1] or ""): r[0] for r in rows if r[1]}

    # MAE_CEST (cesantías)
    cur.execute("SELECT COD_CEST, NOM_CEST FROM MAE_CEST WHERE ACT_ESTA='A'")
    rows = cur.fetchall()
    m["cest_por_nombre"] = {normalize(r[1]): r[0] for r in rows}

    # NO_CONCE
    cur.execute("SELECT COD_CONC, NOM_CONC, TIP_NATU FROM NO_CONCE WHERE ACT_ESTA='A'")
    rows = cur.fetchall()
    m["conc_por_nombre"] = {normalize(r[1]): r[0] for r in rows}
    m["conc_datos"]      = {r[0]: (r[1], r[2]) for r in rows}

    # NO_PERIOD
    cur.execute("SELECT COD_PERIOD, PER_ANO, PER_MES, PER_QNA FROM NO_PERIOD WHERE ACT_ESTA='A'")
    rows = cur.fetchall()
    m["period_por_clave"] = {(r[1], r[2], r[3]): r[0] for r in rows}

    # Empleados existentes en GN_TERCE (NUM_IDEN → COD_TERC)
    cur.execute("SELECT NUM_IDEN, COD_TERC FROM GN_TERCE")
    rows = cur.fetchall()
    m["terc_por_iden"] = {str(r[0]).strip(): r[1] for r in rows if r[0]}

    # GN_FUNCI (COD_TERC → COD_FUNCI)
    cur.execute("SELECT COD_TERC, COD_FUNCI FROM GN_FUNCI")
    rows = cur.fetchall()
    m["funci_por_terc"] = {r[0]: r[1] for r in rows}

    return m


def lookup(diccionario: dict, clave, campo_excel: str, errores: list, fila_ref=None):
    """Busca clave normalizada; agrega error al panel si no existe."""
    key = normalize(str(clave)) if clave else None
    if not key:
        return None
    val = diccionario.get(key)
    if val is None:
        errores.append({
            "fila": fila_ref,
            "campo": campo_excel,
            "valor": clave,
            "mensaje": f"No se encontró '{clave}' en la tabla maestra de {campo_excel}. Agréguelo a la base de datos."
        })
    return val


# ─────────────────────────────────────────────────────────────
# MÓDULO 1 — IMPORTAR EMPLEADOS
# ─────────────────────────────────────────────────────────────

def importar_empleados(wb, conn, cur, mae, errores: list, pendientes: list):
    """
    Lee hoja 'Maestro Original', actualiza/crea registros en GN_TERCE y GN_FUNCI.
    Inhabilita empleados ausentes del Excel.
    """
    print("\n📋  [EMPLEADOS] Procesando hoja 'Maestro Original'...")

    ws = wb["Maestro Original"]
    rows = list(ws.iter_rows(values_only=True))

    # Detectar fila de encabezados (fila 2 → índice 1)
    header_row = rows[1]

    # Mapa de columnas por nombre normalizado
    col_map = {}
    for i, h in enumerate(header_row):
        if h:
            col_map[normalize(str(h))] = i

    def col(nombre):
        return col_map.get(normalize(nombre))

    # Columnas Excel (índices, 0-based)
    IDX = {
        "codigo":         col("Codigo"),
        "codigo_alt":     col("Codigo Alterno"),
        "cedula":         col("Cedula"),
        "tipo_doc":       col("Tipo\nDocumento") or col("Tipo Documento"),
        "nombre":         col("Nombre"),
        "sexo":           col("Sexo"),
        "grupo_san":      col("grupo \nsanguineo") or col("grupo sanguineo"),
        "factor_rh":      col("factor\nrhh") or col("factor rhh"),
        "estado_civ":     col("Estado Civil"),
        "ciu_exped":      col("ciudad expedicion"),
        "hijos":          col("Hijos"),
        "fec_nac":        col("Fecha Nacimiento"),
        "ciudad":         col("Ciudad"),
        "tel1":           col("Telefono1"),
        "tel2":           col("Telefono2"),
        "direccion":      col("Direccion"),
        "correo":         col("Correo"),
        "cargo":          col("Cargo"),
        "valor_hora":     col("Valor Hora"),
        "tipo_cta":       col("Tipo Cuenta"),
        "banco":          col("Banco"),
        "num_cta":        col("Numero Cta"),
        "sucursal":       col("Sucursal"),
        "centro_costo":   col("Centro Costo") or col("Centro costos"),
        "cuenta_gasto":   col("Cuenta de Gasto"),
        "trab_sab":       col("Trabaja Sabado"),
        "clase_sal":      col("Clase Salario"),
        "pensionado":     col("Pensionado"),
        "mod_liquid":     col("Modo Liquidacion"),
        "tip_liquid":     col("Tipo Liquidacion"),
        "extranjero":     col("Extranjero"),
        "reside_ext":     col("Reside Extrnjero"),
        "fec_ingreso":    col("Fecha Ingreso"),
        "fec_retiro":     col("Fecha Retiro"),
        "fec_final":      col("Fecha Final"),
        "cau_retiro":     col("Causa Retiro"),
        "contrato":       col("Contrato"),
        "tip_contrato":   col("Tipo Contrato"),
        "metodo":         col("Metodo"),
        "pct_rete":       col("% Rete"),
        "ded_vivienda":   col("Valor Deduccion 1\nVIVIENDA"),
        "ded_salud":      col("Valor Deduccion2\nSALUD(OTROS)"),
        "ded_depen":      col("Valor Deduccion3\nDEPENDIENTES\n"),
        "pro_salud":      col("Promedio Salud") or col("Promedio salud"),
        "eps":            col("EPS"),
        "afp":            col("AFP"),
        "caja":           col("CAJA"),
        "cesantias":      col("CESANTIAS"),
        "riesgo":         col("Riesgo"),
        "dias_vacac":     col("Dias Vacaciones"),
    }

    cedulas_excel = set()
    insertados = 0
    actualizados = 0

    # Obtener próximos IDs
    cur.execute("SELECT ISNULL(MAX(COD_TERC),0) FROM GN_TERCE")
    next_terc = int(cur.fetchone()[0]) + 1
    cur.execute("SELECT ISNULL(MAX(COD_FUNCI),0) FROM GN_FUNCI")
    next_funci = int(cur.fetchone()[0]) + 1

    # Procesar filas de datos (desde fila 3, índice 2)
    for i, row in enumerate(rows[2:], start=3):
        if not row or all(v is None for v in row):
            continue

        # Cédula (columna Cedula)
        cedula_raw = row[IDX["cedula"]] if IDX["cedula"] is not None else None
        if cedula_raw is None:
            continue
        cedula = str(cedula_raw).strip().split(".")[0]  # quitar decimales
        if not cedula or not cedula.isdigit():
            continue

        cedulas_excel.add(cedula)
        ref = f"Fila {i} ({cedula})"

        def v(key):
            idx = IDX.get(key)
            if idx is None:
                return None
            val = row[idx]
            return val

        # ── Resolver llaves foráneas ──────────────────────────────
        fila_errores_antes = len(errores)

        tipo_doc_str = clean(v("tipo_doc"))
        cod_tpdoc = lookup(mae["tpdoc_por_abrev"], tipo_doc_str, "Tipo Documento (MAE_TPDOC)", errores, ref)

        ciudad_str = clean(v("ciudad"))
        cod_mpio = lookup(mae["muni_por_nombre"], ciudad_str, "Ciudad (MAE_MUNI)", errores, ref) if ciudad_str else None

        ciu_exped_str = clean(v("ciu_exped"))
        cod_ciu_exped = lookup(mae["muni_por_nombre"], ciu_exped_str, "Ciudad Expedición (MAE_MUNI)", errores, ref) if ciu_exped_str else None

        grupo_san_str  = clean(v("grupo_san"))
        factor_rh_str  = clean(v("factor_rh"))
        cod_grsan = None
        if grupo_san_str and factor_rh_str:
            clave_gs = (normalize(grupo_san_str.strip()), normalize(factor_rh_str.strip()))
            cod_grsan = mae["grsan_por_clave"].get(clave_gs)
            if cod_grsan is None:
                errores.append({"fila": ref, "campo": "Grupo Sanguíneo+RH (MAE_GRSAN)",
                                "valor": f"{grupo_san_str}{factor_rh_str}",
                                "mensaje": f"Combinación '{grupo_san_str.strip()}'{factor_rh_str.strip()} no encontrada en MAE_GRSAN."})

        estado_civ_str = clean(v("estado_civ"))
        cod_estciv = lookup(mae["estciv_por_nombre"], estado_civ_str, "Estado Civil (MAE_ESTCIV)", errores, ref) if estado_civ_str else None

        cargo_str = clean(v("cargo"))
        cod_cargo = lookup(mae["cargo_por_nombre"], cargo_str, "Cargo (MAE_CARGO)", errores, ref) if cargo_str else None

        ccost_str = clean(v("centro_costo"))
        cod_ccost = (mae["ccost_por_abrev"].get(normalize(ccost_str))
                     or mae["ccost_por_nombre"].get(normalize(ccost_str))) if ccost_str else None
        if ccost_str and cod_ccost is None:
            errores.append({"fila": ref, "campo": "Centro Costo (MAE_CCOST)",
                            "valor": ccost_str,
                            "mensaje": f"Centro de costo '{ccost_str}' no encontrado en MAE_CCOST."})

        banco_str = clean(v("banco"))
        cod_banco = lookup(mae["banco_por_nombre"], banco_str, "Banco (MAE_BANCO)", errores, ref) if banco_str else None

        tpcta_str = clean(v("tipo_cta"))
        cod_tpcta = (mae["tpcta_por_abrev"].get(normalize(tpcta_str))
                     or mae["tpcta_por_nombre"].get(normalize(tpcta_str))) if tpcta_str else None

        eps_str = clean(v("eps"))
        cod_eps = None
        if eps_str:
            cod_eps = mae["eps_por_nombre"].get(normalize(eps_str))
            if cod_eps is None:
                for k, cv in mae["eps_por_nombre"].items():
                    if k and normalize(eps_str)[:8] in k:
                        cod_eps = cv
                        break
            if cod_eps is None:
                errores.append({"fila": ref, "campo": "EPS (MAE_EPS → GN_TERCE)",
                                "valor": eps_str,
                                "mensaje": f"EPS '{eps_str}' no encontrada. Verifique MAE_EPS / GN_TERCE."})

        afp_str = clean(v("afp"))
        cod_afp = None
        if afp_str:
            # Búsqueda exacta normalizada
            cod_afp = mae["afp_por_nombre"].get(normalize(afp_str))
            # Búsqueda parcial si falla (ej. "COLPENSIONES - I.S.S. PENSION" → "COLPENSIONES")
            if cod_afp is None:
                for k, cv in mae["afp_por_nombre"].items():
                    if k and normalize(afp_str)[:10] in k:
                        cod_afp = cv
                        break
            if cod_afp is None:
                errores.append({"fila": ref, "campo": "AFP (MAE_AFP → GN_TERCE)",
                                "valor": afp_str,
                                "mensaje": f"AFP '{afp_str}' no encontrada. Verifique el nombre en MAE_AFP vía GN_TERCE."})

        caja_str = clean(v("caja"))
        cod_caja = None
        if caja_str:
            cod_caja = mae["ccf_por_nombre"].get(normalize(caja_str))
            if cod_caja is None:
                for k, cv in mae["ccf_por_nombre"].items():
                    if k and normalize(caja_str)[:8] in k:
                        cod_caja = cv
                        break
            if cod_caja is None:
                errores.append({"fila": ref, "campo": "CCF/Caja (MAE_CCF → GN_TERCE)",
                                "valor": caja_str,
                                "mensaje": f"Caja de compensación '{caja_str}' no encontrada en MAE_CCF."})

        cesan_str = clean(v("cesantias"))
        cod_cesan = None
        if cesan_str:
            cod_cesan = mae["cest_por_nombre"].get(normalize(cesan_str))
            if cod_cesan is None:
                # Búsqueda parcial: "PROTECCION CESANTIAS" → "PROTECCION CESANTIAS"
                for k, cv in mae["cest_por_nombre"].items():
                    if k and (normalize(cesan_str)[:8] in k or k[:8] in normalize(cesan_str)):
                        cod_cesan = cv
                        break
            if cod_cesan is None:
                errores.append({"fila": ref, "campo": "Cesantías (MAE_CEST)",
                                "valor": cesan_str,
                                "mensaje": f"Fondo de cesantías '{cesan_str}' no encontrado en MAE_CEST."})

        hay_errores_maestras = len(errores) > fila_errores_antes

        # ── Datos GN_TERCE ────────────────────────────────────────
        nombre_completo = clean(v("nombre"), 240) or ""
        ape_terc, seg_apel, nom_terc, seg_nomb = split_nombre(nombre_completo)

        num_iden_int = int(cedula)
        cod_alt      = clean(v("codigo_alt"), 8)
        tel1         = clean(v("tel1"), 30)
        tel2         = clean(v("tel2"), 40)
        dir_terc     = clean(v("direccion"), 120)
        dir_mail     = clean(v("correo"), 150)

        # ── Datos GN_FUNCI ────────────────────────────────────────
        sexo_raw  = clean(v("sexo"), 1)
        sexo_func = sexo_raw.upper() if sexo_raw else None

        hijos_raw = v("hijos")
        cnt_hijo  = int(hijos_raw) if hijos_raw is not None else 0

        fec_nac      = parse_date(v("fec_nac"))
        fec_ingreso  = parse_date(v("fec_ingreso"))
        fec_retiro_  = parse_date(v("fec_retiro"))
        fec_final    = parse_date(v("fec_final"))

        val_hora_raw = v("valor_hora")
        val_hora     = float(val_hora_raw) if val_hora_raw else None

        num_cta_raw  = v("num_cta")
        num_cta      = int(str(num_cta_raw).strip().replace(" ", "")) if num_cta_raw else None

        sucursal     = clean(v("sucursal"), 10)
        trab_sab     = clean(v("trab_sab"), 10)
        clase_sal    = clean(v("clase_sal"), 10)
        pensionado   = clean(v("pensionado"), 10)
        mod_liquid   = clean(v("mod_liquid"), 10)
        tip_liquid_v = clean(v("tip_liquid"))
        tip_liquid   = None
        if tip_liquid_v:
            try:
                tip_liquid = int(tip_liquid_v)
            except:
                pass

        extranjero   = clean(v("extranjero"), 10)
        reside_ext   = clean(v("reside_ext"), 10)
        cau_retiro   = clean(v("cau_retiro"), 10)
        num_contra_v = v("contrato")
        num_contra   = int(num_contra_v) if num_contra_v else None
        tip_contra   = clean(v("tip_contrato"), 10)
        met_contra   = clean(v("metodo"), 10)
        pct_rete     = clean(v("pct_rete"), 10)

        def to_nchar(val):
            if val is None: return None
            return str(val)[:10]

        ded_viv  = to_nchar(v("ded_vivienda"))
        ded_sal  = to_nchar(v("ded_salud"))
        ded_dep  = to_nchar(v("ded_depen"))
        pro_sal  = to_nchar(v("pro_salud"))
        gra_ries = to_nchar(v("riesgo"))
        dia_vac  = to_nchar(v("dias_vacac"))
        fec_ing_s = fec_ingreso.strftime("%d/%m/%Y") if fec_ingreso else None
        fec_ret_s = fec_retiro_.strftime("%d/%m/%Y") if fec_retiro_ else None
        fec_fin_s = fec_final.strftime("%d/%m/%Y") if fec_final else None

        # ── INSERT o UPDATE GN_TERCE ──────────────────────────────
        cod_terc_existente = mae["terc_por_iden"].get(cedula)

        if hay_errores_maestras:
            pendientes.append({
                "cedula": cedula,
                "nombre": nombre_completo,
                "razon": "Errores en tablas maestras — revisar panel de errores"
            })
            continue

        if cod_terc_existente:
            # UPDATE
            cur.execute("""
                UPDATE GN_TERCE SET
                    COD_ALT   = ?, COD_TPDOC = ?, NOM_COMP  = ?,
                    NOM_TERC  = ?, SEG_NOMB  = ?, APE_TERC  = ?, SEG_APEL = ?,
                    COD_MPIO  = ?, TEL_TERC  = ?, TEL_TERC2 = ?,
                    DIR_TERC  = ?, DIR_MAIL  = ?, ACT_HORA  = GETDATE()
                WHERE COD_TERC = ? AND COD_EMPR = ?
            """, cod_alt, cod_tpdoc, nombre_completo,
                nom_terc[:40], seg_nomb[:40] if seg_nomb else None,
                ape_terc[:40], seg_apel[:40] if seg_apel else None,
                cod_mpio, tel1, tel2, dir_terc, dir_mail,
                cod_terc_existente, COD_EMPR)

            cod_terc = cod_terc_existente
            actualizados += 1
        else:
            # INSERT GN_TERCE
            cur.execute("""
                INSERT INTO GN_TERCE (
                    COD_EMPR, COD_TERC, NUM_IDEN, COD_ALT, COD_TPDOC,
                    NOM_COMP, NOM_TERC, SEG_NOMB, APE_TERC, SEG_APEL,
                    COD_MPIO, TEL_TERC, TEL_TERC2, DIR_TERC, DIR_MAIL,
                    TIP_TERC, TER_EMPL, ACT_INAC, ACT_USUA, ACT_HORA, ACT_ESTA
                ) VALUES (?,?,?,?,?, ?,?,?,?,?, ?,?,?,?,?, ?,?,?,?,GETDATE(),?)
            """, COD_EMPR, next_terc, num_iden_int, cod_alt, cod_tpdoc,
                nombre_completo,
                nom_terc[:40], seg_nomb[:40] if seg_nomb else None,
                ape_terc[:40], seg_apel[:40] if seg_apel else None,
                cod_mpio, tel1, tel2, dir_terc, dir_mail,
                "E", "1", None, "MineDax", "A")

            cod_terc = next_terc
            mae["terc_por_iden"][cedula] = cod_terc
            next_terc += 1
            insertados += 1

        # ── INSERT o UPDATE GN_FUNCI ──────────────────────────────
        cod_funci_existente = mae["funci_por_terc"].get(cod_terc)

        if cod_funci_existente:
            cur.execute("""
                UPDATE GN_FUNCI SET
                    SEX_FUNC=?, COD_GRSAN=?, COD_ESTCIV=?, CNT_HIJO=?,
                    FEC_NAC=?, CIU_EXPED=?, COD_CARGO=?, VAL_HORA=?,
                    COD_TPCTA=?, COD_BANCO=?, NUM_CTA=?, NOM_SUCUR=?,
                    COD_CCOST=?, JOR_SABAD=?, TIP_SALAR=?, CUE_PENSIO=?,
                    MOD_LIQUID=?, TIP_LIQUID=?, EMP_FORAN=?, DIR_FORAN=?,
                    FEC_INGRES=?, FEC_RETIRO=?, FEC_FINAL=?, CAU_RETIRO=?,
                    NUM_CONTRA=?, TIP_CONTRA=?, MET_CONTRA=?, POR_RETEN=?,
                    DED_VIVIEN=?, DED_SALUD=?, DED_DEPEN=?, PRO_SALUD=?,
                    COD_EPS=?, COD_AFP=?, COD_CAJA=?, COD_CESAN=?,
                    GRA_RIESGO=?, DIA_VACAC=?, ACT_HORA=GETDATE()
                WHERE COD_FUNCI=? AND COD_EMPR=?
            """,
                sexo_func, cod_grsan, cod_estciv, cnt_hijo,
                fec_nac, cod_ciu_exped, cod_cargo, val_hora,
                cod_tpcta, cod_banco, num_cta, sucursal,
                cod_ccost, trab_sab, clase_sal, pensionado,
                mod_liquid, tip_liquid, extranjero, reside_ext,
                fec_ing_s, fec_ret_s, fec_fin_s, cau_retiro,
                num_contra, tip_contra, met_contra, pct_rete,
                ded_viv, ded_sal, ded_dep, pro_sal,
                cod_eps, cod_afp, cod_caja, cod_cesan,
                gra_ries, dia_vac,
                cod_funci_existente, COD_EMPR)
        else:
            cur.execute("""
                INSERT INTO GN_FUNCI (
                    COD_EMPR, COD_FUNCI, COD_TERC,
                    SEX_FUNC, COD_GRSAN, COD_ESTCIV, CNT_HIJO,
                    FEC_NAC, CIU_EXPED, COD_CARGO, VAL_HORA,
                    COD_TPCTA, COD_BANCO, NUM_CTA, NOM_SUCUR,
                    COD_CCOST, JOR_SABAD, TIP_SALAR, CUE_PENSIO,
                    MOD_LIQUID, TIP_LIQUID, EMP_FORAN, DIR_FORAN,
                    FEC_INGRES, FEC_RETIRO, FEC_FINAL, CAU_RETIRO,
                    NUM_CONTRA, TIP_CONTRA, MET_CONTRA, POR_RETEN,
                    DED_VIVIEN, DED_SALUD, DED_DEPEN, PRO_SALUD,
                    COD_EPS, COD_AFP, COD_CAJA, COD_CESAN,
                    GRA_RIESGO, DIA_VACAC,
                    ACT_USUA, ACT_HORA, ACT_ESTA
                ) VALUES (
                    ?,?,?,
                    ?,?,?,?,
                    ?,?,?,?,
                    ?,?,?,?,
                    ?,?,?,?,
                    ?,?,?,?,
                    ?,?,?,?,
                    ?,?,?,?,
                    ?,?,?,?,
                    ?,?,?,?,
                    ?,?,
                    ?,GETDATE(),?
                )
            """,
                COD_EMPR, next_funci, cod_terc,
                sexo_func, cod_grsan, cod_estciv, cnt_hijo,
                fec_nac, cod_ciu_exped, cod_cargo, val_hora,
                cod_tpcta, cod_banco, num_cta, sucursal,
                cod_ccost, trab_sab, clase_sal, pensionado,
                mod_liquid, tip_liquid, extranjero, reside_ext,
                fec_ing_s, fec_ret_s, fec_fin_s, cau_retiro,
                num_contra, tip_contra, met_contra, pct_rete,
                ded_viv, ded_sal, ded_dep, pro_sal,
                cod_eps, cod_afp, cod_caja, cod_cesan,
                gra_ries, dia_vac,
                "MineDax", "A")
            mae["funci_por_terc"][cod_terc] = next_funci
            next_funci += 1

    # ── Inhabilitar empleados no presentes en Excel ───────────────
    inhabilitados = 0
    for iden_db, cod_terc_db in list(mae["terc_por_iden"].items()):
        if iden_db not in cedulas_excel:
            # Verificar que sea empleado (TER_EMPL = '1')
            cur.execute("SELECT TER_EMPL FROM GN_TERCE WHERE COD_TERC=?", cod_terc_db)
            row = cur.fetchone()
            if row and str(row[0]).strip() == "1":
                cur.execute("""
                    UPDATE GN_TERCE SET ACT_INAC='I', ACT_HORA=GETDATE()
                    WHERE COD_TERC=? AND COD_EMPR=?
                """, cod_terc_db, COD_EMPR)
                inhabilitados += 1

    print(f"   ✅  Empleados insertados:   {insertados}")
    print(f"   🔄  Empleados actualizados: {actualizados}")
    print(f"   🚫  Empleados inhabilitados:{inhabilitados}")

    conn.commit()


# ─────────────────────────────────────────────────────────────
# MÓDULO 2 — IMPORTAR NOVEDADES
# ─────────────────────────────────────────────────────────────

def get_or_create_period(cur, anio, mes, qna, mae):
    """Retorna COD_PERIOD existente o None si no existe."""
    cod = mae["period_por_clave"].get((anio, mes, qna))
    return cod


def insertar_novedad(cur, cod_funci, cod_conc, cod_period, obs, mae) -> Optional[int]:
    """
    Inserta en NO_NOVED. Si ya existe el mismo (empleado+concepto+período), retorna el existente.
    """
    cur.execute("""
        SELECT COD_NOVED FROM NO_NOVED
        WHERE COD_FUNCI=? AND COD_CONC=? AND COD_PERIOD=? AND COD_EMPR=?
    """, cod_funci, cod_conc, cod_period, COD_EMPR)
    row = cur.fetchone()
    if row:
        return row[0]

    cur.execute("SELECT ISNULL(MAX(COD_NOVED),0)+1 FROM NO_NOVED")
    next_id = cur.fetchone()[0]

    cur.execute("""
        INSERT INTO NO_NOVED (COD_EMPR, COD_NOVED, COD_FUNCI, COD_CONC, COD_PERIOD,
                              FEC_REGI, OBS_NOVED, IND_APLICADO, ACT_USUA, ACT_HORA, ACT_ESTA)
        VALUES (?,?,?,?,?, CONVERT(date,GETDATE()),?,?,?,GETDATE(),?)
    """, COD_EMPR, next_id, cod_funci, cod_conc, cod_period,
        obs or None, "N", "MineDax", "A")
    return next_id


def importar_ocasionales(wb, cur, mae, anio, mes, qna, errores):
    print("\n📊  [NOVEDADES] Procesando hoja 'Ocasionales'...")
    ws  = wb["Ocasionales"]
    rows = list(ws.iter_rows(values_only=True))

    cod_period = get_or_create_period(cur, anio, mes, qna, mae)
    if cod_period is None:
        errores.append({"fila": "Ocasionales", "campo": "Período",
                        "valor": f"{anio}-{mes}-Q{qna}",
                        "mensaje": f"El período {anio}/{mes} Q{qna} no existe en NO_PERIOD."})
        return 0

    ins = 0
    for i, row in enumerate(rows[8:], start=9):  # datos desde fila 9
        if not row or row[1] is None:
            continue
        cedula   = str(row[1]).strip().split(".")[0]
        novedad  = clean(row[3])
        cantidad = row[5]
        valor    = row[6]
        obs      = clean(row[7], 500)

        if not cedula or not cedula.isdigit() or not novedad:
            continue

        cod_terc  = mae["terc_por_iden"].get(cedula)
        cod_funci = mae["funci_por_terc"].get(cod_terc) if cod_terc else None
        if not cod_funci:
            errores.append({"fila": f"Ocasionales fila {i}", "campo": "Empleado",
                            "valor": cedula,
                            "mensaje": f"Cédula {cedula} no encontrada en la base de datos."})
            continue

        cod_conc = mae["conc_por_nombre"].get(normalize(novedad))
        if cod_conc is None:
            errores.append({"fila": f"Ocasionales fila {i}", "campo": "Concepto",
                            "valor": novedad,
                            "mensaje": f"Concepto '{novedad}' no encontrado en NO_CONCE."})
            continue

        cod_noved = insertar_novedad(cur, cod_funci, cod_conc, cod_period, obs, mae)

        # NO_OCASI — actualiza si ya existe, inserta si no
        cur.execute("SELECT COD_NOVED FROM NO_OCASI WHERE COD_NOVED=? AND COD_EMPR=?",
                    cod_noved, COD_EMPR)
        if cur.fetchone():
            cur.execute("UPDATE NO_OCASI SET CANTIDAD=?, VALOR=?, ACT_HORA=GETDATE() WHERE COD_NOVED=? AND COD_EMPR=?",
                        cantidad, valor, cod_noved, COD_EMPR)
        else:
            cur.execute("""
                INSERT INTO NO_OCASI (COD_EMPR, COD_NOVED, CANTIDAD, VALOR, ACT_USUA, ACT_HORA, ACT_ESTA)
                VALUES (?,?,?,?,?,GETDATE(),?)
            """, COD_EMPR, cod_noved, cantidad, valor, "MineDax", "A")
            ins += 1

    print(f"   ✅  Ocasionales procesadas: {ins} nuevas")
    return ins


def importar_fijas(wb, cur, mae, anio, mes, qna, errores):
    print("\n📌  [NOVEDADES] Procesando hoja 'Fijas'...")
    ws   = wb["Fijas"]
    rows = list(ws.iter_rows(values_only=True))

    cod_period = get_or_create_period(cur, anio, mes, qna, mae)
    if cod_period is None:
        errores.append({"fila": "Fijas", "campo": "Período",
                        "valor": f"{anio}-{mes}-Q{qna}",
                        "mensaje": f"El período {anio}/{mes} Q{qna} no existe en NO_PERIOD."})
        return 0

    ins = 0
    for i, row in enumerate(rows[8:], start=9):
        if not row or row[1] is None:
            continue
        cedula      = str(row[1]).strip().split(".")[0]
        novedad     = clean(row[3])
        cantidad    = row[5]
        valor       = row[6]
        fec_ini     = parse_date(row[7])
        fec_fin     = parse_date(row[8])
        aplicacion  = clean(row[9], 30)
        num_cuenta  = clean(row[10], 50)
        cuotas      = row[11]
        obs         = clean(row[12], 500)

        if not cedula or not cedula.isdigit() or not novedad:
            continue

        cod_terc  = mae["terc_por_iden"].get(cedula)
        cod_funci = mae["funci_por_terc"].get(cod_terc) if cod_terc else None
        if not cod_funci:
            errores.append({"fila": f"Fijas fila {i}", "campo": "Empleado",
                            "valor": cedula,
                            "mensaje": f"Cédula {cedula} no encontrada en la base de datos."})
            continue

        cod_conc = mae["conc_por_nombre"].get(normalize(novedad))
        if cod_conc is None:
            errores.append({"fila": f"Fijas fila {i}", "campo": "Concepto",
                            "valor": novedad,
                            "mensaje": f"Concepto '{novedad}' no encontrado en NO_CONCE."})
            continue

        cod_noved = insertar_novedad(cur, cod_funci, cod_conc, cod_period, obs, mae)

        cur.execute("SELECT COD_NOVED FROM NO_FIJAS WHERE COD_NOVED=? AND COD_EMPR=?",
                    cod_noved, COD_EMPR)
        if cur.fetchone():
            cur.execute("""
                UPDATE NO_FIJAS SET CANTIDAD=?, VALOR=?, FEC_INI=?, FEC_FIN=?,
                    APLICACION=?, NUM_CUOTAS=?, NUM_CUENTA=?, ACT_HORA=GETDATE()
                WHERE COD_NOVED=? AND COD_EMPR=?
            """, cantidad, valor, fec_ini, fec_fin,
                aplicacion, cuotas, num_cuenta,
                cod_noved, COD_EMPR)
        else:
            cur.execute("""
                INSERT INTO NO_FIJAS (COD_EMPR, COD_NOVED, CANTIDAD, VALOR,
                    FEC_INI, FEC_FIN, APLICACION, NUM_CUOTAS, NUM_CUENTA,
                    ACT_USUA, ACT_HORA, ACT_ESTA)
                VALUES (?,?,?,?,?,?,?,?,?,?,GETDATE(),?)
            """, COD_EMPR, cod_noved, cantidad, valor,
                fec_ini, fec_fin, aplicacion, cuotas, num_cuenta,
                "MineDax", "A")
            ins += 1

    print(f"   ✅  Fijas procesadas: {ins} nuevas")
    return ins


def importar_ausentismos(wb, cur, mae, anio, mes, qna, errores):
    print("\n🏥  [NOVEDADES] Procesando hoja 'Ausentismos Vacaciones'...")
    ws   = wb["Ausentismos Vacaciones"]
    rows = list(ws.iter_rows(values_only=True))

    cod_period = get_or_create_period(cur, anio, mes, qna, mae)
    if cod_period is None:
        errores.append({"fila": "Ausentismos", "campo": "Período",
                        "valor": f"{anio}-{mes}-Q{qna}",
                        "mensaje": f"El período {anio}/{mes} Q{qna} no existe en NO_PERIOD."})
        return 0

    ins = 0
    for i, row in enumerate(rows[9:], start=10):  # encabezado en fila 9 (índice 8)
        if not row or row[1] is None:
            continue
        cedula      = str(row[1]).strip().split(".")[0]
        ausentismo  = clean(row[3])
        fec_ini     = parse_date(row[4])
        fec_fin     = parse_date(row[5])
        dias        = row[6]
        diagnostico = clean(row[7], 20)
        prorroga    = clean(row[8])
        obs         = clean(row[9], 500)

        if not cedula or not cedula.isdigit() or not ausentismo:
            continue

        cod_terc  = mae["terc_por_iden"].get(cedula)
        cod_funci = mae["funci_por_terc"].get(cod_terc) if cod_terc else None
        if not cod_funci:
            errores.append({"fila": f"Ausentismos fila {i}", "campo": "Empleado",
                            "valor": cedula,
                            "mensaje": f"Cédula {cedula} no encontrada en la base de datos."})
            continue

        cod_conc = mae["conc_por_nombre"].get(normalize(ausentismo))
        if cod_conc is None:
            errores.append({"fila": f"Ausentismos fila {i}", "campo": "Concepto",
                            "valor": ausentismo,
                            "mensaje": f"Tipo de ausentismo '{ausentismo}' no encontrado en NO_CONCE."})
            continue

        cod_noved = insertar_novedad(cur, cod_funci, cod_conc, cod_period, obs, mae)

        cur.execute("SELECT COD_NOVED FROM NO_AUSEN WHERE COD_NOVED=? AND COD_EMPR=?",
                    cod_noved, COD_EMPR)
        if cur.fetchone():
            cur.execute("""
                UPDATE NO_AUSEN SET FEC_INI=?, FEC_FIN=?, DIAS_TOTAL=?,
                    DIAGNOSTICO=?, FEC_PRORRG=?, ACT_HORA=GETDATE()
                WHERE COD_NOVED=? AND COD_EMPR=?
            """, fec_ini, fec_fin, dias, diagnostico,
                parse_date(prorroga) if prorroga and prorroga.lower() not in ("si","no","sí") else None,
                cod_noved, COD_EMPR)
        else:
            cur.execute("""
                INSERT INTO NO_AUSEN (COD_EMPR, COD_NOVED, FEC_INI, FEC_FIN, DIAS_TOTAL,
                    DIAGNOSTICO, FEC_PRORRG, ACT_USUA, ACT_HORA, ACT_ESTA)
                VALUES (?,?,?,?,?,?,?,?,GETDATE(),?)
            """, COD_EMPR, cod_noved, fec_ini, fec_fin, dias,
                diagnostico,
                parse_date(prorroga) if prorroga and prorroga.lower() not in ("si","no","sí") else None,
                "MineDax", "A")
            ins += 1

    print(f"   ✅  Ausentismos procesados: {ins} nuevos")
    return ins


def importar_cambios(wb, cur, mae, anio, mes, qna, errores):
    print("\n🔄  [NOVEDADES] Procesando hoja 'Cambios e Ingresos'...")
    ws   = wb["Cambios e Ingresos"]
    rows = list(ws.iter_rows(values_only=True))

    cod_period = get_or_create_period(cur, anio, mes, qna, mae)
    if cod_period is None:
        errores.append({"fila": "Cambios", "campo": "Período",
                        "valor": f"{anio}-{mes}-Q{qna}",
                        "mensaje": f"El período {anio}/{mes} Q{qna} no existe en NO_PERIOD."})
        return 0

    ins = 0
    for i, row in enumerate(rows[8:], start=9):
        if not row or row[1] is None:
            continue
        cedula     = str(row[1]).strip().split(".")[0]
        cambio     = clean(row[3])
        fec_ini    = parse_date(row[4])
        valor_nuevo = clean(row[5], 300)
        obs        = clean(row[6], 500)

        if not cedula or not cedula.isdigit() or not cambio:
            continue

        cod_terc  = mae["terc_por_iden"].get(cedula)
        cod_funci = mae["funci_por_terc"].get(cod_terc) if cod_terc else None
        if not cod_funci:
            errores.append({"fila": f"Cambios fila {i}", "campo": "Empleado",
                            "valor": cedula,
                            "mensaje": f"Cédula {cedula} no encontrada en la base de datos."})
            continue

        cod_conc = mae["conc_por_nombre"].get(normalize(cambio))
        if cod_conc is None:
            errores.append({"fila": f"Cambios fila {i}", "campo": "Concepto",
                            "valor": cambio,
                            "mensaje": f"Tipo de cambio '{cambio}' no encontrado en NO_CONCE."})
            continue

        cod_noved = insertar_novedad(cur, cod_funci, cod_conc, cod_period, obs, mae)

        cur.execute("SELECT COD_NOVED FROM NO_CAMBI WHERE COD_NOVED=? AND COD_EMPR=?",
                    cod_noved, COD_EMPR)
        if cur.fetchone():
            cur.execute("""
                UPDATE NO_CAMBI SET FEC_INI=?, VALOR_NUEVO=?, ACT_HORA=GETDATE()
                WHERE COD_NOVED=? AND COD_EMPR=?
            """, fec_ini, valor_nuevo, cod_noved, COD_EMPR)
        else:
            cur.execute("""
                INSERT INTO NO_CAMBI (COD_EMPR, COD_NOVED, FEC_INI, VALOR_NUEVO, VALOR_ANTE,
                    ACT_USUA, ACT_HORA, ACT_ESTA)
                VALUES (?,?,?,?,?,?,GETDATE(),?)
            """, COD_EMPR, cod_noved, fec_ini, valor_nuevo, None, "MineDax", "A")
            ins += 1

    print(f"   ✅  Cambios procesados: {ins} nuevos")
    return ins


# ─────────────────────────────────────────────────────────────
# MENÚ INTERACTIVO
# ─────────────────────────────────────────────────────────────

def mostrar_menu():
    print("\n" + "═" * 60)
    print("   IMPORTADOR ADECCO / COLLECTIVE MINING → MineDax")
    print("═" * 60)
    print("   Este archivo contiene dos tipos de información:\n")
    print("   [1]  Importar solo EMPLEADOS")
    print("        (Actualiza GN_TERCE y GN_FUNCI desde 'Maestro Original')")
    print()
    print("   [2]  Importar solo NOVEDADES")
    print("        (Ocasionales, Fijas, Ausentismos, Cambios)")
    print()
    print("   [3]  Importar AMBOS (empleados primero, luego novedades)")
    print()
    print("   [0]  Cancelar")
    print("─" * 60)

    while True:
        try:
            opc = input("   Seleccione una opción [0-3]: ").strip()
            if opc in ("0", "1", "2", "3"):
                return int(opc)
            print("   ⚠️  Opción inválida. Intente de nuevo.")
        except (KeyboardInterrupt, EOFError):
            return 0


def imprimir_panel_errores(errores: list, pendientes: list):
    print("\n" + "═" * 60)
    print("   PANEL DE ERRORES")
    print("═" * 60)
    if not errores:
        print("   ✅  Sin errores de validación.")
    else:
        print(f"   ⚠️  Se encontraron {len(errores)} error(es):\n")
        for e in errores:
            print(f"   📍 {e['fila']}  |  {e['campo']}")
            print(f"      Valor: {e['valor']}")
            print(f"      {e['mensaje']}")
            print()

    if pendientes:
        print(f"\n   ⏸️  Registros pendientes (no importados por errores): {len(pendientes)}")
        for p in pendientes:
            print(f"      • {p['cedula']} — {p['nombre']}: {p['razon']}")

    print("═" * 60)


# ─────────────────────────────────────────────────────────────
# PUNTO DE ENTRADA
# ─────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Uso: python importar_adecco.py <ruta_excel>")
        sys.exit(1)

    ruta = sys.argv[1]
    if not os.path.exists(ruta):
        print(f"❌  Archivo no encontrado: {ruta}")
        sys.exit(1)

    print(f"\n📂  Archivo: {os.path.basename(ruta)}")

    # Detectar período desde nombre del archivo
    anio, mes, qna = extraer_periodo_nombre(ruta)
    if anio:
        print(f"📅  Período detectado: {anio} / Mes {mes} / Quincena {qna}")
    else:
        print("⚠️  No se pudo detectar el período desde el nombre del archivo.")

    # Menú
    opcion = mostrar_menu()
    if opcion == 0:
        print("   Operación cancelada.")
        return

    hacer_empleados = opcion in (1, 3)
    hacer_novedades = opcion in (2, 3)

    # Cargar Excel (desencripta en memoria si tiene contraseña)
    print("\n📖  Cargando archivo Excel...")
    try:
        wb = abrir_excel(ruta)
    except RuntimeError as e:
        print(f"\n❌  {e}")
        sys.exit(1)

    # Conectar DB
    print("🔌  Conectando a MineDax...")
    conn = get_conn()
    cur  = conn.cursor()

    errores   = []
    pendientes = []

    try:
        # Cargar maestras
        print("🗂️   Cargando tablas maestras...")
        mae = cargar_maestras(cur)

        # Prioridad: empleados primero (para que novedades encuentren los nuevos)
        if hacer_empleados:
            importar_empleados(wb, conn, cur, mae, errores, pendientes)

        if hacer_novedades:
            if anio is None:
                print("\n❌  No se puede importar novedades sin período. Cancele e ingrese un archivo con el período en el nombre (ej: 2Q_FEB_2026).")
            else:
                importar_ocasionales(wb, cur, mae, anio, mes, qna, errores)
                importar_fijas(wb, cur, mae, anio, mes, qna, errores)
                importar_ausentismos(wb, cur, mae, anio, mes, qna, errores)
                importar_cambios(wb, cur, mae, anio, mes, qna, errores)

        conn.commit()
        print("\n✅  Importación completada y guardada en la base de datos.")

    except Exception as ex:
        conn.rollback()
        print(f"\n❌  Error inesperado: {ex}")
        import traceback
        traceback.print_exc()
    finally:
        imprimir_panel_errores(errores, pendientes)
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
