#!/usr/bin/env python3
# =============================================================================
#  scripts/generar_adecco.py
#
#  Recibe los datos de novedades por stdin (JSON, UTF-8) y genera un Excel
#  basado en la plantilla FORMATO_LIBRE_ADECCO.xlsx.
#  Escribe el archivo resultante en la ruta indicada por argv[1].
#
#  Estructura JSON esperada en stdin:
#  {
#    "periodo":      { "etiqueta": "2026 - Abril - Q1" },
#    "ocasionales":  [ { IDENTIFICACION, NOMBRE, NOVEDAD, TIPO_NOVEDAD, CANTIDAD, VALOR, OBSERVACIONES } ],
#    "fijas":        [ { IDENTIFICACION, NOMBRE, NOVEDAD, TIPO_NOVEDAD, CANTIDAD, VALOR, FEC_INI, FEC_FIN, APLICACION, CUENTA, CUOTAS, OBSERVACIONES } ],
#    "ausentismos":  [ { IDENTIFICACION, NOMBRE, AUSENTISMO, FECHA_INICIAL, FECHA_FINAL, DIAS_TOTALES, DIAGNOSTICO, PRORROGA, OBSERVACIONES } ],
#    "cambios":      [ { IDENTIFICACION, NOMBRE, CAMBIO, FECHA_INICIAL, CAMBIO_A, OBSERVACIONES } ]
#  }
#
#  ENCODING: stdin y stdout se fuerzan a UTF-8 explícitamente para garantizar
#  el correcto manejo de tildes, ñ, apóstrofes y caracteres de lenguas romance
#  y anglosajonas en cualquier plataforma (Windows, Linux, macOS).
# =============================================================================

import sys
import io
import json
import shutil
import os
from copy import copy
from datetime import datetime

# ─── Forzar UTF-8 en stdin/stdout/stderr ─────────────────────────────────────
# En Windows el encoding por defecto puede ser cp1252 o similar, lo que
# corrompe caracteres como á, é, í, ó, ú, ñ, ü, â, ê, apostrofes tipográficos, etc.
# Forzamos UTF-8 independientemente del entorno.
sys.stdin  = io.TextIOWrapper(sys.stdin.buffer,  encoding='utf-8', errors='replace')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

OUTPUT_PATH      = sys.argv[1]
TEMPLATE_PATH    = os.path.join(os.path.dirname(__file__), '..', 'assets', 'FORMATO_LIBRE_ADECCO.xlsx')
MAESTRO_REF_PATH = os.path.join(os.path.dirname(__file__), '..', 'assets', 'ultimo_maestro_adecco.xlsx')

# ─── Leer datos desde stdin (ya forzado a UTF-8) ─────────────────────────────
data = json.load(sys.stdin)

ocasionales = data.get('ocasionales', [])
fijas       = data.get('fijas',       [])
ausentismos = data.get('ausentismos', [])
cambios     = data.get('cambios',     [])
etiqueta    = data.get('periodo', {}).get('etiqueta', '')

# ─── Copiar plantilla ─────────────────────────────────────────────────────────
shutil.copy2(TEMPLATE_PATH, OUTPUT_PATH)
wb = openpyxl.load_workbook(OUTPUT_PATH)

# ─── Helpers ─────────────────────────────────────────────────────────────────

def copy_sheet(src_ws, dst_wb, sheet_name):
    """
    Copia src_ws hacia dst_wb bajo sheet_name, reemplazando si ya existe.
    Preserva valores, estilos, dimensiones de columnas/filas y celdas fusionadas.
    """
    # Recordar posición actual en el workbook destino
    pos = None
    if sheet_name in dst_wb.sheetnames:
        pos = dst_wb.sheetnames.index(sheet_name)
        del dst_wb[sheet_name]

    dst_ws = dst_wb.create_sheet(title=sheet_name)

    # Copiar valores y estilos celda a celda
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst_cell.font          = copy(cell.font)
                dst_cell.fill          = copy(cell.fill)
                dst_cell.border        = copy(cell.border)
                dst_cell.alignment     = copy(cell.alignment)
                dst_cell.number_format = cell.number_format

    # Dimensiones de columnas
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width  = col_dim.width
        dst_ws.column_dimensions[col_letter].hidden = col_dim.hidden

    # Dimensiones de filas
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height
        dst_ws.row_dimensions[row_num].hidden = row_dim.hidden

    # Celdas fusionadas
    for merge_range in list(src_ws.merged_cells.ranges):
        dst_ws.merge_cells(str(merge_range))

    # Restaurar posición original si la hoja ya existía
    if pos is not None:
        dst_wb.move_sheet(dst_ws, offset=pos - (len(dst_wb.sheetnames) - 1))

    return dst_ws


def parse_date(val):
    """Convierte string ISO o None a datetime.date (sin hora)."""
    if not val:
        return None
    try:
        return datetime.fromisoformat(str(val).replace('Z', '+00:00')).date()
    except Exception:
        return None

def copy_row_style(ws, src_row_idx, dst_row_idx):
    """Copia estilos (fill, font, alignment, border, number_format) de src a dst."""
    for col_idx in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row_idx, column=col_idx)
        dst_cell = ws.cell(row=dst_row_idx, column=col_idx)
        if src_cell.has_style:
            dst_cell.font        = copy(src_cell.font)
            dst_cell.fill        = copy(src_cell.fill)
            dst_cell.border      = copy(src_cell.border)
            dst_cell.alignment   = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format

def clear_data_rows(ws, first_data_row):
    """Borra valores (no estilos) de todas las filas de datos existentes."""
    for row_idx in range(first_data_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).value = None

def write_rows(ws, first_data_row, rows, col_map):
    """
    Escribe filas de datos en la hoja.
    col_map = lista de (column_letter_or_index, field_name, type)
    type: 'str' | 'num' | 'date'
    """
    # Usar fila de estilo de referencia (primera fila de datos en la plantilla)
    style_ref_row = first_data_row

    for i, record in enumerate(rows):
        row_idx = first_data_row + i

        # Copiar estilo de la fila de referencia si existe en la plantilla
        if row_idx > style_ref_row:
            copy_row_style(ws, style_ref_row, row_idx)

        for col_spec in col_map:
            col_idx, field, tipo = col_spec
            val = record.get(field)

            cell = ws.cell(row=row_idx, column=col_idx)

            if val is None or val == '':
                cell.value = None
                continue

            if tipo == 'num':
                try:
                    cell.value = float(val)
                    if cell.value == int(cell.value):
                        cell.value = int(cell.value)
                except (ValueError, TypeError):
                    cell.value = val
            elif tipo == 'date':
                d = parse_date(val)
                cell.value = d
                if d:
                    cell.number_format = 'DD/MM/YYYY'
            else:
                cell.value = str(val)


# =============================================================================
#  OCASIONALES
#  Encabezados en fila 8: B=Identificación C=Nombre D=Novedad E=Tipo F=cantidad G=valor H=observaciones
#  Datos desde fila 9
# =============================================================================
ws_oc = wb['Ocasionales']
FIRST_ROW_OC = 9

clear_data_rows(ws_oc, FIRST_ROW_OC)

col_map_oc = [
    (2, 'IDENTIFICACION', 'str'),   # B
    (3, 'NOMBRE',         'str'),   # C
    (4, 'NOVEDAD',        'str'),   # D
    (5, 'TIPO_NOVEDAD',   'str'),   # E
    (6, 'CANTIDAD',       'num'),   # F
    (7, 'VALOR',          'num'),   # G
    (8, 'OBSERVACIONES',  'str'),   # H
]
write_rows(ws_oc, FIRST_ROW_OC, ocasionales, col_map_oc)


# =============================================================================
#  FIJAS
#  Encabezados en fila 8: B=Identificación C=Nombre D=Novedad E=Tipo F=cantidad G=valor
#                          H=Fecha Inicial I=Fecha Final J=Aplicación K=Cuenta L=Cuotas M=observaciones
#  Datos desde fila 9
# =============================================================================
ws_fj = wb['Fijas']
FIRST_ROW_FJ = 9

clear_data_rows(ws_fj, FIRST_ROW_FJ)

col_map_fj = [
    (2,  'IDENTIFICACION', 'str'),   # B
    (3,  'NOMBRE',         'str'),   # C
    (4,  'NOVEDAD',        'str'),   # D
    (5,  'TIPO_NOVEDAD',   'str'),   # E
    (6,  'CANTIDAD',       'num'),   # F
    (7,  'VALOR',          'num'),   # G
    (8,  'FEC_INI',        'date'),  # H
    (9,  'FEC_FIN',        'date'),  # I
    (10, 'APLICACION',     'str'),   # J
    (11, 'CUENTA',         'str'),   # K
    (12, 'CUOTAS',         'num'),   # L
    (13, 'OBSERVACIONES',  'str'),   # M
]
write_rows(ws_fj, FIRST_ROW_FJ, fijas, col_map_fj)


# =============================================================================
#  AUSENTISMOS VACACIONES
#  Encabezados en fila 9: B=Identificación C=Nombre D=Ausentismo E=Fecha Inicial F=Fecha Final
#                          G=Dias totales H=diagnostico I=Prorroga J=observaciones
#  Datos desde fila 10
# =============================================================================
ws_au = wb['Ausentismos Vacaciones']
FIRST_ROW_AU = 10

clear_data_rows(ws_au, FIRST_ROW_AU)

col_map_au = [
    (2,  'IDENTIFICACION', 'str'),   # B
    (3,  'NOMBRE',         'str'),   # C
    (4,  'AUSENTISMO',     'str'),   # D
    (5,  'FECHA_INICIAL',  'date'),  # E
    (6,  'FECHA_FINAL',    'date'),  # F
    (7,  'DIAS_TOTALES',   'num'),   # G
    (8,  'DIAGNOSTICO',    'str'),   # H
    (9,  'PRORROGA',       'date'),  # I
    (10, 'OBSERVACIONES',  'str'),   # J
]
write_rows(ws_au, FIRST_ROW_AU, ausentismos, col_map_au)


# =============================================================================
#  CAMBIOS E INGRESOS
#  Encabezados en fila 8: B=Identificación C=Nombre D=Cambio E=Fecha Inicial F=cambio a G=observaciones
#  Datos desde fila 9
# =============================================================================
ws_ci = wb['Cambios e Ingresos']
FIRST_ROW_CI = 9

clear_data_rows(ws_ci, FIRST_ROW_CI)

col_map_ci = [
    (2, 'IDENTIFICACION', 'str'),   # B
    (3, 'NOMBRE',         'str'),   # C
    (4, 'CAMBIO',         'str'),   # D
    (5, 'FECHA_INICIAL',  'date'),  # E
    (6, 'CAMBIO_A',       'str'),   # F
    (7, 'OBSERVACIONES',  'str'),   # G
]
write_rows(ws_ci, FIRST_ROW_CI, cambios, col_map_ci)


# ─── Reemplazar hojas maestras desde el último ADECCO importado ───────────────
# Si existe un archivo de referencia guardado al importar (ultimo_maestro_adecco.xlsx),
# se copian sus hojas "Maestro Original" y "Cambios Maestro" al workbook de salida.
# Así el exportado refleja siempre la plantilla maestra más reciente, no la estática.
if os.path.exists(MAESTRO_REF_PATH):
    try:
        ref_wb = openpyxl.load_workbook(MAESTRO_REF_PATH, data_only=True)
        for nombre_hoja in ('Maestro Original', 'Cambios Maestro'):
            if nombre_hoja in ref_wb.sheetnames:
                copy_sheet(ref_wb[nombre_hoja], wb, nombre_hoja)
        ref_wb.close()
    except Exception as e:
        print(f'AVISO: No se pudieron copiar las hojas maestras de referencia: {e}', file=sys.stderr)

# ─── Guardar ─────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f'OK:{OUTPUT_PATH}', flush=True)
